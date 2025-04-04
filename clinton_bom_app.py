import streamlit as st
import pandas as pd

# --- Data ---
# Dictionary holding the details for each Clinton part we know about
clinton_parts = {
    # Part Num: {desc: Description, type: 'pole' or 'accessory', cost: unit price}
    "CE-CP6W": {"desc": "Telescoping Pole w/Bracket, Ceiling Mount, 6ft Adjustable, Aluminum/Steel, White", "type": "pole", "cost": 29.47},
    "CE-CP6B": {"desc": "Telescoping Pole w/Bracket, Ceiling Mount, 6ft Adjustable, Aluminum/Steel, Black", "type": "pole", "cost": 29.47},
    "CE-CP3W": {"desc": "Telescoping Pole w/Bracket, Ceiling Mount, 3ft Adjustable, Aluminum/Steel, White", "type": "pole", "cost": 24.49},
    "CE-CP3B": {"desc": "Telescoping Pole w/Bracket, Ceiling Mount, 3ft Adjustable, Aluminum/Steel, Black", "type": "pole", "cost": 24.49},
    "CE-CP12W": {"desc": "Telescoping Pole w/Bracket, Ceiling Mount, 12ft Adjustable, Aluminum/Steel, White", "type": "pole", "cost": 35.52},
    "CE-CP12B": {"desc": "Telescoping Pole w/Bracket, Ceiling Mount, 12ft Adjustable, Aluminum/Steel, Black", "type": "pole", "cost": 35.52},
    "CE-CP17W": {"desc": "Telescoping Pole w/Bracket, Ceiling Mount, 6FT-17FT Adjustable, Aluminum/Steel, White", "type": "pole", "cost": 68.63},
    "CE-CP17B": {"desc": "Telescoping Pole w/Bracket, Ceiling Mount, 6FT-17FT Adjustable, Aluminum/Steel, Black", "type": "pole", "cost": 68.63},
    "CE-CPUP": {"desc": "UNIVERSAL MOUNTING PLATE FOR TELESCOPING CAMERA POLES", "type": "accessory", "cost": 9.59},
    "CE-CPBCM": {"desc": "Camera Pole Beam Clamp", "type": "accessory", "cost": 12.44},
    # Add more parts here if needed following the same format
}

# Dictionary for material parts
material_parts = {
    "10136230": {"desc": "White CAT6 CAT6 Cable 1000' BOX", "manufacturer": "Nexxt Inc", "supplier": "Graybar", "cost": 234.71},
    "NK688MBU": {"desc": "Blue Cat6 Jack", "manufacturer": "Panduit", "supplier": "Graybar", "cost": 5.85},
    "NK2BXWH-A": {"desc": "2 Port SMB", "manufacturer": "Panduit", "supplier": "Graybar", "cost": 11.75},
    "INFINI CAB CAT6-01WH": {"desc": "1' White Cat6 Patch Cord", "manufacturer": "Infinite", "supplier": "Graybar", "cost": 2.95},
    "AT1610-WH": {"desc": "16 White Cat5 Patch Cord", "manufacturer": "Allen Tel", "supplier": "Graybar", "cost": 6.60},
    "NK77M": {"desc": "24 port cat6 patch panel", "manufacturer": "Panduit", "supplier": "Graybar", "cost": 151.84},
    "1000PC": {"desc": "CAT6e RJ45 8P8C Modular Plugs (50 pc)", "manufacturer": "Platinum", "supplier": "Graybar", "cost": 39.90},
    "11081": {"desc": "Velcro", "manufacturer": "Platinum", "supplier": "Graybar", "cost": 29.90},
    "HW-AC-03": {"desc": "3/4\" x 1 in. Zinc Plated Steel Flat Head Phillips/Square Machine Screws (pack of 100)", "manufacturer": "E. F. Cable", "supplier": "Graybar", "cost": 14.75},
    "FHP20MM24UM": {"desc": "2\" Hook Runner On Hinge Clip", "manufacturer": "E. F. Cable", "supplier": "Graybar", "cost": 7.50}
}

# Get just the part numbers that are poles for creating the input fields
pole_part_nums = sorted([p for p, d in clinton_parts.items() if d["type"] == "pole"])

# --- BoM Generation Function ---
def generate_clinton_bom(project_id, reader_count, pole_quantities):
    """
    Generates the Bill of Materials list based on user inputs.
    Args:
        project_id (str): The project identifier.
        reader_count (int): Total number of readers (determines accessory counts).
        pole_quantities (dict): Dictionary with pole part numbers as keys and quantities as values.
    Returns:
        list: A list of dictionaries, where each dictionary represents a line item in the BoM.
    """
    bom_items = []
    supplier = "Clinton"
    manufacturer = "Clinton" # Assuming Clinton is always the manufacturer for these parts
    price_expiration = "12/31/2025" # Setting a default price expiration date

    # Process Poles
    for part_num, qty in pole_quantities.items():
        # Only add if quantity is positive and it's a known part
        if qty > 0 and part_num in clinton_parts:
            part_info = clinton_parts[part_num]
            cost = part_info.get("cost", 0)
            extended_cost = cost * qty
            bom_items.append({
                "Project": project_id,
                "Required Supplier": supplier,
                "Manufacturer": manufacturer,
                "Manufacturer Part #": part_num,
                "Description": part_info["desc"],
                "Quantity": qty,
                "Cost": cost,
                "Extended Cost": extended_cost,
                "Price Expiration": price_expiration
            })

    # Calculate the total number of poles
    total_poles = sum(qty for part, qty in pole_quantities.items() if clinton_parts[part]["type"] == "pole")

    # Process Accessories (Plates and Clamps) based on pole count instead of reader count
    if total_poles > 0:
        # Mounting Plates
        part_num_plate = "CE-CPUP"
        if part_num_plate in clinton_parts:
             part_info_plate = clinton_parts[part_num_plate]
             cost = part_info_plate.get("cost", 0)
             extended_cost = cost * total_poles
             bom_items.append({
                "Project": project_id,
                "Required Supplier": supplier,
                "Manufacturer": manufacturer,
                "Manufacturer Part #": part_num_plate,
                "Description": part_info_plate["desc"],
                "Quantity": total_poles,
                "Cost": cost,
                "Extended Cost": extended_cost,
                "Price Expiration": price_expiration
            })
        # Beam Clamps
        part_num_clamp = "CE-CPBCM"
        if part_num_clamp in clinton_parts:
            part_info_clamp = clinton_parts[part_num_clamp]
            cost = part_info_clamp.get("cost", 0)
            extended_cost = cost * total_poles
            bom_items.append({
                "Project": project_id,
                "Required Supplier": supplier,
                "Manufacturer": manufacturer,
                "Manufacturer Part #": part_num_clamp,
                "Description": part_info_clamp["desc"],
                "Quantity": total_poles,
                "Cost": cost,
                "Extended Cost": extended_cost,
                "Price Expiration": price_expiration
            })

    return bom_items

# Function to generate material BoM
def generate_material_bom(project_id, reader_count, cable_quantity):
    """
    Generates the Material Bill of Materials list based on user inputs.
    Args:
        project_id (str): The project identifier.
        reader_count (int): Total number of readers (determines quantity for several items).
        cable_quantity (int): Quantity of network cable boxes.
    Returns:
        list: A list of dictionaries, where each dictionary represents a line item in the BoM.
    """
    bom_items = []
    price_expiration = "12/31/2025"  # Setting a default price expiration date

    # Process all materials
    for part_num, part_info in material_parts.items():
        qty = 1  # Default quantity
        
        # Special rules for quantities based on requirements
        if part_num == "10136230":  # Cable box
            qty = cable_quantity
        elif part_num == "NK688MBU":  # Blue Cat6 Jack
            qty = reader_count * 2
        elif part_num in ["NK2BXWH-A", "INFINI CAB CAT6-01WH", "AT1610-WH"]:
            qty = reader_count
            
        # Only add parts with positive quantities
        if qty > 0:
            cost = part_info.get("cost", 0)
            extended_cost = cost * qty
            bom_items.append({
                "Project": project_id,
                "Required Supplier": part_info["supplier"],
                "Manufacturer": part_info["manufacturer"],
                "Manufacturer Part #": part_num,
                "Description": part_info["desc"],
                "Quantity": qty,
                "Cost": cost,
                "Extended Cost": extended_cost,
                "Price Expiration": price_expiration
            })

    return bom_items

# --- Streamlit App Layout ---

st.set_page_config(layout="wide") # Use wider layout for more space
st.title("RFID BoM Generator")
st.write("Fill in the details below to generate the Bill of Materials for Clinton poles and material orders.")

# Create tabs for Clinton Poles and Material Order
tab1, tab2 = st.tabs(["Clinton Pole Order", "Material Order"])

# --- Session State for Sharing Data Between Tabs ---
if 'project_id' not in st.session_state:
    st.session_state.project_id = ""
if 'store_name' not in st.session_state:
    st.session_state.store_name = ""
if 'reader_count' not in st.session_state:
    st.session_state.reader_count = 0
if 'cable_quantity' not in st.session_state:
    st.session_state.cable_quantity = 1

# ----- CLINTON POLE ORDER TAB -----
with tab1:
    st.header("Clinton Pole Order")
    
    # --- Input Section ---
    st.subheader("1. Enter Order Details")

    # Use columns for better layout
    col1, col2 = st.columns([1, 2]) # Make second column wider

    with col1:
        project_id = st.text_input("Project ID:", key="project_id_clinton", value=st.session_state.project_id)
        store_name = st.text_input("Store Name:", key="store_name_clinton", value=st.session_state.store_name)
        reader_count = st.number_input("Total Number of Readers:", min_value=0, step=1, value=st.session_state.reader_count, key="reader_count_clinton",
                                       help="This determines the quantity for Mounting Plates (CE-CPUP) and Beam Clamps (CE-CPBCM)")
        
        # Update session state when values change
        st.session_state.project_id = project_id
        st.session_state.store_name = store_name
        st.session_state.reader_count = reader_count

    with col2:
        st.subheader("Pole Quantities")
        pole_quantities_input = {}
        
        # Group poles by height and color for better organization
        # 3ft poles
        st.markdown("**1-3ft Poles**")
        if "CE-CP3W" in clinton_parts:
            label = "CE-CP3W (1-3ft White Pole)"
            pole_quantities_input["CE-CP3W"] = st.number_input(label, min_value=0, step=1, value=0, key=f"qty_CE-CP3W")
        if "CE-CP3B" in clinton_parts:
            label = "CE-CP3B (1-3ft Black Pole)"
            pole_quantities_input["CE-CP3B"] = st.number_input(label, min_value=0, step=1, value=0, key=f"qty_CE-CP3B")
        
        # 6ft poles
        st.markdown("**3-6ft Poles**")
        if "CE-CP6W" in clinton_parts:
            label = "CE-CP6W (3-6ft White Pole)"
            pole_quantities_input["CE-CP6W"] = st.number_input(label, min_value=0, step=1, value=0, key=f"qty_CE-CP6W")
        if "CE-CP6B" in clinton_parts:
            label = "CE-CP6B (3-6ft Black Pole)"
            pole_quantities_input["CE-CP6B"] = st.number_input(label, min_value=0, step=1, value=0, key=f"qty_CE-CP6B")
        
        # 12ft poles
        st.markdown("**6-12ft Poles**")
        if "CE-CP12W" in clinton_parts:
            label = "CE-CP12W (6-12ft White Pole)"
            pole_quantities_input["CE-CP12W"] = st.number_input(label, min_value=0, step=1, value=0, key=f"qty_CE-CP12W")
        if "CE-CP12B" in clinton_parts:
            label = "CE-CP12B (6-12ft Black Pole)"
            pole_quantities_input["CE-CP12B"] = st.number_input(label, min_value=0, step=1, value=0, key=f"qty_CE-CP12B")
        
        # 17ft poles
        st.markdown("**17ft Poles**")
        if "CE-CP17W" in clinton_parts:
            label = "CE-CP17W (17ft White Pole)"
            pole_quantities_input["CE-CP17W"] = st.number_input(label, min_value=0, step=1, value=0, key=f"qty_CE-CP17W")
        if "CE-CP17B" in clinton_parts:
            label = "CE-CP17B (17ft Black Pole)"
            pole_quantities_input["CE-CP17B"] = st.number_input(label, min_value=0, step=1, value=0, key=f"qty_CE-CP17B")
        
        # Make sure we include any other poles that might be added in the future
        other_poles = [p for p in pole_part_nums if p not in ["CE-CP3W", "CE-CP3B", "CE-CP6W", "CE-CP6B", "CE-CP12W", "CE-CP12B", "CE-CP17W", "CE-CP17B"]]
        if other_poles:
            st.markdown("**Other Poles**")
            for part_num in other_poles:
                # Extract a short description part for the label if possible
                try:
                    short_desc = clinton_parts[part_num]['desc'].split(',')[1].strip() + ", " + clinton_parts[part_num]['desc'].split(',')[-1].strip()
                except:
                    short_desc = clinton_parts[part_num]['desc'][:30] + "..." # Fallback short description

                label = f"{part_num} ({short_desc})"
                pole_quantities_input[part_num] = st.number_input(label, min_value=0, step=1, value=0, key=f"qty_{part_num}")

    # --- Generate BoM Button and Output Section ---
    st.header("2. Generate Bill of Materials")

    if st.button("Generate BoM", key="generate_button"):
        # Basic validation
        if not project_id:
            st.warning("⚠️ Please enter a Project ID.")
            st.stop() # Stop execution if no project ID

        if reader_count == 0 and all(qty == 0 for qty in pole_quantities_input.values()):
             st.warning("⚠️ Please enter quantities for at least one pole OR set 'Total Number of Readers' > 0.")
             st.stop() # Stop execution if no quantities are entered

        # Call the generation function
        generated_bom_list = generate_clinton_bom(project_id, reader_count, pole_quantities_input)

        if generated_bom_list:
            # Convert list of dictionaries to Pandas DataFrame for better display & CSV export
            bom_df = pd.DataFrame(generated_bom_list)

            # Ensure columns are in the desired order
            column_order = ["Project", "Required Supplier", "Manufacturer", "Manufacturer Part #", 
                            "Description", "Quantity", "Cost", "Extended Cost", "Price Expiration"]
            # Make sure all expected columns exist, add if missing (shouldn't happen with current logic)
            for col in column_order:
                 if col not in bom_df.columns:
                     bom_df[col] = None # Add missing column if needed
            bom_df = bom_df[column_order] # Reorder

            # Make a display copy with formatted currency
            display_df = bom_df.copy()
            display_df["Cost"] = display_df["Cost"].map('${:.2f}'.format)
            display_df["Extended Cost"] = display_df["Extended Cost"].map('${:.2f}'.format)

            st.subheader("Generated Clinton BoM")
            st.dataframe(display_df, hide_index=True, use_container_width=True) # Display the table

            # --- Export to CSV ---
            # Convert DataFrame to CSV format in memory
            csv = bom_df.to_csv(index=False).encode('utf-8')
            
            # --- Export to Excel ---
            # Function to create an Excel file in memory
            def create_excel_bytes(df, store_name):
                import io
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.utils.dataframe import dataframe_to_rows
                
                # Create a workbook and select the active worksheet
                wb = Workbook()
                ws = wb.active
                
                # Limit sheet title to 31 characters to avoid openpyxl warning
                safe_sheet_title = "Clinton BoM"
                ws.title = safe_sheet_title
                
                # Define styles
                header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                header_font = Font(bold=True)
                border = Border(
                    left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin')
                )
                
                # Add dataframe data to worksheet - except Extended Cost which will be formula-based
                rows = dataframe_to_rows(df, index=False, header=True)
                
                # Keep track of the last data row for total calculation
                last_data_row = 0
                
                for r_idx, row in enumerate(rows, 1):
                    for c_idx, value in enumerate(row, 1):
                        # Skip Extended Cost column (8) for data rows - will add formula later
                        if c_idx == 8 and r_idx > 1:
                            continue
                        
                        cell = ws.cell(row=r_idx, column=c_idx, value=value)
                        
                        # Add border to all cells
                        cell.border = border
                        
                        # Style header row
                        if r_idx == 1:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = Alignment(horizontal='center')
                        
                        # Format currency columns
                        if c_idx == 7 and r_idx > 1:  # Cost column
                            cell.number_format = '$#,##0.00'
                
                # Keep track of the last data row
                if r_idx > 1:
                    last_data_row = r_idx
            
                # Add formulas for Extended Cost column
                for row in range(2, last_data_row + 1):
                    cell = ws.cell(row=row, column=8)
                    cell.value = f"=F{row}*G{row}"  # Quantity * Cost
                    cell.number_format = '$#,##0.00'
                    cell.border = border
                
                # Add total row
                total_row = last_data_row + 1
                
                # Add the word "Total" in the second-to-last column of the total row
                total_label_cell = ws.cell(row=total_row, column=7)
                total_label_cell.value = "Total"
                total_label_cell.font = Font(bold=True)
                total_label_cell.border = border
                
                # Add the SUM formula for the Extended Cost column
                sum_cell = ws.cell(row=total_row, column=8)
                if last_data_row > 1:  # Only add sum if there's data
                    sum_cell.value = f"=SUM(H2:H{last_data_row})"
                else:
                    sum_cell.value = 0
                sum_cell.number_format = '$#,##0.00'
                sum_cell.font = Font(bold=True)
                sum_cell.border = border
                
                # Adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        if cell.value:
                            try:
                                cell_length = len(str(cell.value))
                                if cell_length > max_length:
                                    max_length = cell_length
                            except:
                                pass
                
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
                # Save to memory buffer
                buffer = io.BytesIO()
                wb.save(buffer)
                buffer.seek(0)
                return buffer
        
        # Create the Excel file
        excel_bytes = create_excel_bytes(bom_df, store_name)
        
        # Create a download buttons row
        col1, col2 = st.columns(2)
        
        with col1:
            st.download_button(
                label="Download BoM as CSV",
                data=csv,
                # Sanitize project_id for filename
                file_name=f'{project_id.replace(" ", "_").replace("/", "-")}_Clinton_BoM.csv',
                mime='text/csv',
                key='download_csv'
            )
            
        with col2:
            filename = f'Clinton BOM {store_name}'
            if not store_name:
                filename = f'Clinton BOM {project_id.replace(" ", "_").replace("/", "-")}'
                
            st.download_button(
                label="Download BoM as Excel",
                data=excel_bytes,
                file_name=f'{filename}.xlsx',
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                key='download_excel'
            )
            
        # Store in session state so it can be potentially redisplayed or reused
        st.session_state['generated_bom_df'] = bom_df
        st.session_state['project_id_generated'] = project_id # Store project ID for filename consistency
    else:
        st.info("ℹ️ No BoM items were generated based on the input provided.")
        # Clear any old BoM from session state
        if 'generated_bom_df' in st.session_state:
            del st.session_state['generated_bom_df']
            del st.session_state['project_id_generated']

# ----- MATERIAL ORDER TAB -----
with tab2:
    st.header("Material Order")
    
    # --- Input Section ---
    st.subheader("1. Enter Material Order Details")
    
    # Use columns for better layout
    col1, col2 = st.columns([1, 2])
    
    with col1:
        # Use the existing project ID and store name from session state
        st.write(f"**Project ID:** {st.session_state.project_id}")
        st.write(f"**Store Name:** {st.session_state.store_name}")
        st.write(f"**Reader Count:** {st.session_state.reader_count}")
        
        # Additional input for cable quantity
        cable_quantity = st.number_input("Network Cable Boxes (1000' each):", min_value=1, step=1, 
                                         value=st.session_state.cable_quantity, key="cable_qty")
        st.session_state.cable_quantity = cable_quantity
    
    with col2:
        st.subheader("Material Quantities")
        st.write("The following quantities will be automatically calculated:")
        st.markdown(f"- **Blue Cat6 Jack (NK688MBU):** {st.session_state.reader_count * 2} (Reader Count × 2)")
        st.markdown(f"- **2 Port SMB (NK2BXWH-A):** {st.session_state.reader_count} (Reader Count)")
        st.markdown(f"- **1' White Cat6 Patch Cord (INFINI CAB CAT6-01WH):** {st.session_state.reader_count} (Reader Count)")
        st.markdown(f"- **16 White Cat5 Patch Cord (AT1610-WH):** {st.session_state.reader_count} (Reader Count)")
        st.markdown("- All other materials: Quantity of 1")
    
    # --- Generate Material BoM Button and Output Section ---
    st.header("2. Generate Material Bill of Materials")
    
    if st.button("Generate Material BoM", key="generate_material_button"):
        # Basic validation
        if not st.session_state.project_id:
            st.warning("⚠️ Please enter a Project ID in the Clinton Pole Order tab.")
            st.stop()
        
        if st.session_state.reader_count == 0:
            st.warning("⚠️ Please enter a Reader Count greater than 0 in the Clinton Pole Order tab.")
            st.stop()
        
        # Call the generation function
        generated_bom_list = generate_material_bom(st.session_state.project_id, 
                                                  st.session_state.reader_count, 
                                                  cable_quantity)
        
        if generated_bom_list:
            # Convert list of dictionaries to Pandas DataFrame for better display & Excel export
            bom_df = pd.DataFrame(generated_bom_list)
            
            # Ensure columns are in the desired order
            column_order = ["Project", "Required Supplier", "Manufacturer", "Manufacturer Part #", 
                            "Description", "Quantity", "Cost", "Extended Cost", "Price Expiration"]
            # Make sure all expected columns exist, add if missing
            for col in column_order:
                if col not in bom_df.columns:
                    bom_df[col] = None
            bom_df = bom_df[column_order]  # Reorder
            
            # Make a display copy with formatted currency
            display_df = bom_df.copy()
            display_df["Cost"] = display_df["Cost"].map('${:.2f}'.format)
            display_df["Extended Cost"] = display_df["Extended Cost"].map('${:.2f}'.format)
            
            st.subheader("Generated Material BoM")
            st.dataframe(display_df, hide_index=True, use_container_width=True)
            
            # --- Export to Excel ---
            # Function to create an Excel file in memory (same as in Clinton tab)
            def create_excel_bytes(df, store_name):
                import io
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.utils.dataframe import dataframe_to_rows
                
                # Create a workbook and select the active worksheet
                wb = Workbook()
                ws = wb.active
                
                # Limit sheet title to 31 characters to avoid openpyxl warning
                safe_sheet_title = "Material BoM"
                ws.title = safe_sheet_title
                
                # Define styles
                header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                header_font = Font(bold=True)
                border = Border(
                    left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin')
                )
                
                # Add dataframe data to worksheet - except Extended Cost which will be formula-based
                rows = dataframe_to_rows(df, index=False, header=True)
                
                # Keep track of the last data row for total calculation
                last_data_row = 0
                
                for r_idx, row in enumerate(rows, 1):
                    for c_idx, value in enumerate(row, 1):
                        # Skip Extended Cost column (8) for data rows - will add formula later
                        if c_idx == 8 and r_idx > 1:
                            continue
                            
                        cell = ws.cell(row=r_idx, column=c_idx, value=value)
                        
                        # Add border to all cells
                        cell.border = border
                        
                        # Style header row
                        if r_idx == 1:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = Alignment(horizontal='center')
                        
                        # Format currency columns
                        if c_idx == 7 and r_idx > 1:  # Cost column
                            cell.number_format = '$#,##0.00'
                    
                    # Keep track of the last data row
                    if r_idx > 1:
                        last_data_row = r_idx
                
                # Add formulas for Extended Cost column
                for row in range(2, last_data_row + 1):
                    cell = ws.cell(row=row, column=8)
                    cell.value = f"=F{row}*G{row}"  # Quantity * Cost
                    cell.number_format = '$#,##0.00'
                    cell.border = border
                
                # Add total row
                total_row = last_data_row + 1
                
                # Add the word "Total" in the second-to-last column of the total row
                total_label_cell = ws.cell(row=total_row, column=7)
                total_label_cell.value = "Total"
                total_label_cell.font = Font(bold=True)
                total_label_cell.border = border
                
                # Add the SUM formula for the Extended Cost column
                sum_cell = ws.cell(row=total_row, column=8)
                if last_data_row > 1:  # Only add sum if there's data
                    sum_cell.value = f"=SUM(H2:H{last_data_row})"
                else:
                    sum_cell.value = 0
                sum_cell.number_format = '$#,##0.00'
                sum_cell.font = Font(bold=True)
                sum_cell.border = border
                
                # Adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        if cell.value:
                            try:
                                cell_length = len(str(cell.value))
                                if cell_length > max_length:
                                    max_length = cell_length
                            except:
                                pass
                    
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Save to memory buffer
                buffer = io.BytesIO()
                wb.save(buffer)
                buffer.seek(0)
                return buffer
            
            # Create the Excel file
            excel_bytes = create_excel_bytes(bom_df, st.session_state.store_name)
            
            # Determine filename based on store name
            filename = f'RFID Material Order {st.session_state.store_name}'
            if not st.session_state.store_name:
                filename = f'RFID Material Order {st.session_state.project_id.replace(" ", "_").replace("/", "-")}'
            
            # Download button
            st.download_button(
                label="Download Material BoM as Excel",
                data=excel_bytes,
                file_name=f'{filename}.xlsx',
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                key='download_material_excel'
            )
            
            # Store in session state
            st.session_state['generated_material_bom_df'] = bom_df
        else:
            st.info("ℹ️ No Material BoM items were generated.")

# --- How to Run ---
st.sidebar.header("How to Run")
st.sidebar.markdown("""
1.  **Save:** Save this code as a Python file (e.g., `clinton_bom_app.py`).
2.  **Open Terminal:** Open your command prompt or terminal.
3.  **Navigate:** Go to the directory where you saved the file.
    ```bash
    cd path/to/your/directory
    ```
4.  **Run:** Execute the Streamlit command:
    ```bash
    streamlit run clinton_bom_app.py
    ```
5.  **Use:** Your web browser should automatically open with the tool running. If not, the terminal will provide a local URL (usually `http://localhost:8501`).
""")
st.sidebar.header("Add More Parts")
st.sidebar.markdown("""
To add more Clinton pole types or accessories:
1. Edit the `clinton_parts = {...}` dictionary near the top of the script.
2. Add a new entry following the format:
   `"PART-NUMBER": {"desc": "Full Description", "type": "pole" (or "accessory"), "cost": unit price}`
3. If it's a `pole`, it will automatically appear in the 'Pole Quantities' input section when you rerun the app. Accessories are handled automatically based on the reader count.
""")
