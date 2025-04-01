import openpyxl
from openpyxl.styles import PatternFill

# Load the workbook
wb = openpyxl.load_workbook('Clinton BOM GRG 387 - Extra Pack.xlsx')
ws = wb.active

print(f"Sheet name: {ws.title}")
print(f"Total rows: {ws.max_row}")
print(f"Total columns: {ws.max_column}")

# Print column widths
print("\nColumn widths:")
for col_letter in 'ABCDEFGHIJ':
    if col_letter in ws.column_dimensions:
        width = ws.column_dimensions[col_letter].width
        print(f"  Column {col_letter}: {width}")

# Print header information
print("\nHeader row (row 1):")
for cell in list(ws.rows)[0]:
    if cell.value:
        print(f"  {cell.coordinate}: '{cell.value}'")
        print(f"    Font: {cell.font.name}, Size: {cell.font.size}, Bold: {cell.font.bold}")
        if isinstance(cell.fill, PatternFill) and cell.fill.start_color.rgb:
            print(f"    Fill color: {cell.fill.start_color.rgb}")
        print(f"    Alignment: horizontal={cell.alignment.horizontal}, vertical={cell.alignment.vertical}")
        print(f"    Format: {cell.number_format}")

# Check for formulas in the Extended Cost column
print("\nChecking for formulas in Extended Cost column:")
for row_idx in range(2, min(6, ws.max_row + 1)):  # Check first few data rows
    cell = ws.cell(row=row_idx, column=8)  # Assuming Extended Cost is column H (8)
    print(f"  Cell H{row_idx}: value={cell.value}, formula={getattr(cell, 'formula', 'None')}")

# Check the total row
total_row = ws.max_row
print(f"\nTotal row (row {total_row}):")
for col_idx in range(1, ws.max_column + 1):
    cell = ws.cell(row=total_row, column=col_idx)
    if cell.value:
        print(f"  Cell {cell.coordinate}: '{cell.value}'")
        print(f"    Font: Bold: {cell.font.bold}")
        print(f"    Format: {cell.number_format}")
        print(f"    Formula: {getattr(cell, 'formula', 'None')}")
