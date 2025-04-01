# <file name="rfid_bom_generator.py">
import streamlit as st
import pandas as pd
import io
import re
import os
from datetime import date
from fuzzywuzzy import process
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Color
from openpyxl.utils.dataframe import dataframe_to_rows

# --- Data ---
# Dictionary holding the details for each Clinton part we know about
clinton_parts = {
    # Part Num: {desc: Description, type: 'pole' or 'accessory', cost: unit price}
    "CE-CP8B": {"desc": "8' Fixed Height Pole, Black", "type": "pole", "cost": 31.36},
    "CE-CP6W": {"desc": "Telescoping Pole w/Bracket, Ceiling Mount, 6ft Adjustable, Aluminum/Steel, White", "type": "pole", "cost": 29.47},
    "CE-CP6B": {"desc": "Telescoping Pole w/Bracket, Ceiling Mount, 6ft Adjustable, Aluminum/Steel, Black", "type": "pole", "cost": 29.47},
    "CE-CP3W": {"desc": "Telescoping Pole w/Bracket, Ceiling Mount, 3ft Adjustable, Aluminum/Steel, White", "type": "pole", "cost": 24.49},
    "CE-CP3B": {"desc": "Telescoping Pole w/Bracket, Ceiling Mount, 3ft Adjustable, Aluminum/Steel, Black", "type": "pole", "cost": 24.49},
    "CE-CP412B": {"desc": "Adjustable from 3' 11.25\" to 10' 11.25\", Black, UL", "type": "pole", "cost": 35.52},
    "CE-CP412B-2PK": {"desc": "Adjustable from 3' 11.25\" to 10' 11.25\", Black, UL Two Poles per Box, Sold as Pair", "type": "pole", "cost": 63.36, "is_pair": True},
    "CE-CP412W": {"desc": "Adjustable from 3' 11.25\" to 10' 11.25\", White, UL", "type": "pole", "cost": 35.52},
    "CE-CP412W-2PK": {"desc": "Adjustable from 3' 11.25\" to 10' 11.25\", White, UL Two Poles per Box, Sold as Pair", "type": "pole", "cost": 63.36, "is_pair": True},
    "CE-CP16B": {"desc": "16' Fixed Height Pole, Black", "type": "pole", "cost": 63.36},
    "CE-CP20B": {"desc": "20' Fixed Height Pole, Black", "type": "pole", "cost": 79.20},
    # Accessories
    "CE-CPUP": {"desc": "UNIVERSAL MOUNTING PLATE FOR TELESCOPING CAMERA POLES", "type": "accessory", "cost": 9.59},
    "CE-CPBCM": {"desc": "Camera Pole Beam Clamp", "type": "accessory", "cost": 12.44},
    # Add more parts here if needed following the same format
}

# Dictionary for material parts (Using Anixter data from the first image)
material_parts = {
    "10136230": {"desc": "White CMP CAT6 Cable 1000' BOX", "manufacturer": "Berk-Tek", "supplier": "Anixter", "cost": 264.94},
    "NK688MBU": {"desc": "Blue Cat6 Jack", "manufacturer": "Panduit", "supplier": "Anixter", "cost": 5.88},
    "NK2BXWH-A": {"desc": "2 Port SMB", "manufacturer": "Panduit", "supplier": "Anixter", "cost": 1.70},
    "INFINI CAB CAT6-01WH": {"desc": "1' White Cat6 Patch Cord", "manufacturer": "INFINI", "supplier": "Anixter", "cost": 2.05},
    "AT1610-WH": {"desc": "10' White Cat6 Patch Cord", "manufacturer": "Allen Tel", "supplier": "Anixter", "cost": 8.60},
    "NKFP24Y": {"desc": "24-port cat6 patch panel", "manufacturer": "Panduit", "supplier": "Anixter", "cost": 31.84},
    "100003C": {"desc": "CAT5e RJ-45 8P8C Modular Plugs (50 pk)", "manufacturer": "Platinum", "supplier": "Anixter", "cost": 35.00},
    "31086": {"desc": "Velcro", "manufacturer": "Velcro", "supplier": "Anixter", "cost": 23.00},
    "FMDD6321": {"desc": "6-32 x 1 in. Zinc Plated Steel Flat Head Phillips/Square Machine Screws (pack of 100)", "manufacturer": "L. H. Dottie", "supplier": "Anixter", "cost": 4.76},
    "CAT32HP24SM": {"desc": "2\" j-hook Hammer On Flange Clip", "manufacturer": "B-line", "supplier": "Anixter", "cost": 7.50}
}


# Get just the part numbers that are poles for creating the input fields
pole_part_nums = sorted([p for p, d in clinton_parts.items() if d["type"] == "pole"])

# --- BoM Generation Functions ---
def generate_clinton_bom(project_id, store_name, reader_count, pole_quantities):
    """
    Generate Bill of Materials (BoM) for Clinton poles and accessories.
    
    Args:
        project_id (str): Project identifier.
        store_name (str): Store name.
        reader_count (int): Number of readers.
        pole_quantities (dict): Dictionary with pole part numbers as keys and quantities as values.
    
    Returns:
        list: List of dictionaries containing BoM items data.
    """
    bom_items = []
    supplier = "Clinton" # Required supplier - always Clinton
    manufacturer = "Clinton" # Assuming Clinton is always the manufacturer for these parts

    # Process the special case for the black adjustable poles (CE-CP412B and CE-CP412B-2PK)
    # If there are 2 or more CE-CP412B poles, use the 2-pack instead
    if "CE-CP412B" in pole_quantities and pole_quantities["CE-CP412B"] >= 2:
        # Calculate how many 2-packs we need
        pairs_needed = pole_quantities["CE-CP412B"] // 2
        singles_needed = pole_quantities["CE-CP412B"] % 2
        
        # Add the 2-packs to the BoM
        if pairs_needed > 0:
            bom_items.append({
                "Project": project_id,
                "Required Supplier": supplier,
                "Manufacturer": manufacturer,
                "Manufacturer Part #": "CE-CP412B-2PK",
                "Description": clinton_parts["CE-CP412B-2PK"]["desc"],
                "Quantity": pairs_needed,
                "Cost": clinton_parts["CE-CP412B-2PK"]["cost"]
            })
        
        # Add any remaining single poles
        if singles_needed > 0:
            bom_items.append({
                "Project": project_id,
                "Required Supplier": supplier,
                "Manufacturer": manufacturer,
                "Manufacturer Part #": "CE-CP412B",
                "Description": clinton_parts["CE-CP412B"]["desc"],
                "Quantity": singles_needed,
                "Cost": clinton_parts["CE-CP412B"]["cost"]
            })
        
        # Remove the processed poles from the dictionary
        del pole_quantities["CE-CP412B"]
    
    # Process the special case for the white adjustable poles (CE-CP412W and CE-CP412W-2PK)
    # If there are 2 or more CE-CP412W poles, use the 2-pack instead
    if "CE-CP412W" in pole_quantities and pole_quantities["CE-CP412W"] >= 2:
        # Calculate how many 2-packs we need
        pairs_needed = pole_quantities["CE-CP412W"] // 2
        singles_needed = pole_quantities["CE-CP412W"] % 2
        
        # Add the 2-packs to the BoM
        if pairs_needed > 0:
            bom_items.append({
                "Project": project_id,
                "Required Supplier": supplier,
                "Manufacturer": manufacturer,
                "Manufacturer Part #": "CE-CP412W-2PK",
                "Description": clinton_parts["CE-CP412W-2PK"]["desc"],
                "Quantity": pairs_needed,
                "Cost": clinton_parts["CE-CP412W-2PK"]["cost"]
            })
        
        # Add any remaining single poles
        if singles_needed > 0:
            bom_items.append({
                "Project": project_id,
                "Required Supplier": supplier,
                "Manufacturer": manufacturer,
                "Manufacturer Part #": "CE-CP412W",
                "Description": clinton_parts["CE-CP412W"]["desc"],
                "Quantity": singles_needed,
                "Cost": clinton_parts["CE-CP412W"]["cost"]
            })
        
        # Remove the processed poles from the dictionary
        del pole_quantities["CE-CP412W"]
    
    # Add remaining poles to BoM
    for part_num, qty in pole_quantities.items():
        if qty > 0 and part_num in clinton_parts:
            part_info = clinton_parts[part_num]
            
            # Handle the case where this is a 2-pack directly selected by the user
            if part_num in ["CE-CP412B-2PK", "CE-CP412W-2PK"]:
                # Each 2-pack counts as 2 poles, but we add it as-is since the user specifically selected it
                bom_items.append({
                    "Project": project_id,
                    "Required Supplier": supplier,
                    "Manufacturer": manufacturer,
                    "Manufacturer Part #": part_num,
                    "Description": part_info["desc"],
                    "Quantity": qty,
                    "Cost": part_info["cost"]
                })
                continue
            
            # Regular case for all other poles
            bom_items.append({
                "Project": project_id,
                "Required Supplier": supplier,
                "Manufacturer": manufacturer,
                "Manufacturer Part #": part_num,
                "Description": part_info["desc"],
                "Quantity": qty,
                "Cost": part_info["cost"]
            })
    
    # Add accessories based on total pole quantity
    # Calculate total number of poles, counting 2-packs as 2 poles each
    total_poles = 0
    
    # Count regular poles first
    for part_num, qty in pole_quantities.items():
        if qty > 0:
            if part_num.endswith("-2PK"):
                # Each 2-pack counts as 2 poles
                total_poles += qty * 2
            else:
                total_poles += qty
    
    # Count poles from items already added to the BoM (needed for 2-packs)
    for item in bom_items:
        part_num = item["Manufacturer Part #"]
        # Skip non-pole items
        if part_num not in clinton_parts or clinton_parts[part_num]["type"] != "pole":
            continue
            
        if part_num.endswith("-2PK"):
            # Each 2-pack counts as 2 poles
            total_poles += item["Quantity"] * 2
        else:
            total_poles += item["Quantity"]
    
    # Add necessary accessories
    if total_poles > 0:
        # Add universal mounting plates equal to the number of POLES
        bom_items.append({
            "Project": project_id,
            "Required Supplier": supplier,
            "Manufacturer": manufacturer,
            "Manufacturer Part #": "CE-CPUP",
            "Description": clinton_parts["CE-CPUP"]["desc"],
            "Quantity": total_poles,  
            "Cost": clinton_parts["CE-CPUP"]["cost"]
        })
        
        # Add beam clamps equal to the number of POLES
        bom_items.append({
            "Project": project_id,
            "Required Supplier": supplier,
            "Manufacturer": manufacturer,
            "Manufacturer Part #": "CE-CPBCM",
            "Description": clinton_parts["CE-CPBCM"]["desc"],
            "Quantity": total_poles,  
            "Cost": clinton_parts["CE-CPBCM"]["cost"]
        })
    
    return bom_items

def generate_material_bom(project_id, reader_count, cable_quantity):
    """
    Generates the Material Bill of Materials list based on user inputs.
    Args:
        project_id (str): The project identifier.
        reader_count (int): Total number of readers (determines quantity for several items).
        cable_quantity (int): Quantity of network cable boxes.
    Returns:
        list: A list of dictionaries, where each dictionary represents a line item in the BoM.
    """
    bom_items = []
    # price_expiration = "12/31/2025"  # Removed, handled by formatting

    # Process all materials
    for part_num, part_info in material_parts.items():
        qty = 1  # Default quantity

        # Special rules for quantities based on requirements
        if part_num == "10136230":  # Cable box
            qty = cable_quantity
        elif part_num == "NK688MBU":  # Blue Cat6 Jack
            qty = reader_count * 2
        elif part_num in ["NK2BXWH-A", "INFINI CAB CAT6-01WH", "AT1610-WH"]:
            qty = reader_count

        # Only add parts with positive quantities
        if qty > 0:
            cost = part_info.get("cost", 0)
            # extended_cost = cost * qty # Calculated by Excel formula
            bom_items.append({
                "Project": project_id,
                "Required Supplier": part_info["supplier"],
                "Manufacturer": part_info["manufacturer"],
                "Manufacturer Part #": part_num,
                "Description": part_info["desc"],
                "Quantity": qty,
                "Cost": cost,
                # "Extended Cost": extended_cost, # Removed
                # "Price Expiration": price_expiration # Removed
            })

    return bom_items

# Helper function to create Excel files
def create_excel_bytes(df, sheet_title, store_name=None):
    """
    Creates an Excel file with styling and formulas precisely matching the target.
    Args:
        df (pandas.DataFrame): DataFrame with the BoM data (must contain columns A-G).
        sheet_title (str): Title for the worksheet.
        store_name (str, optional): Store name for reference.
    Returns:
        BytesIO: Excel file in memory.
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Color
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    # Limit sheet title to 31 characters to avoid openpyxl warning
    safe_sheet_title = sheet_title[:31] if len(sheet_title) > 31 else sheet_title
    
    # Create a workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = safe_sheet_title

    # --- Define Styles to EXACTLY match the reference Excel file ---
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
    ext_cost_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")  # Light green

    header_font = Font(name='Arial', size=10, bold=True, color="000000")
    normal_font = Font(name='Arial', size=10, bold=False, color="000000")

    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
    right_alignment = Alignment(horizontal='right', vertical='center', wrap_text=False)

    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    # Currency format with $ sign fixed left
    currency_format = '"$"#,##0.00'
    # Integer format for quantity
    integer_format = '0'

    # --- Set column widths to match the reference file exactly ---
    # Obtained by inspecting the target Excel file
    column_widths = {
        'A': 11.89, # Project (Adjusted slightly from Graybar example for PRJ...)
        'B': 15.0,  # Required Supplier
        'C': 14.0,  # Manufacturer
        'D': 18.0,  # Manufacturer Part #
        'E': 60.0,  # Description (Adjusted slightly)
        'F': 8.0,   # Quantity
        'G': 10.0,  # Cost
        'H': 14.0,  # Extended Cost
        'I': 8.0,   # Empty column / TOTAL label
        'J': 15.0   # Price Expiration
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # --- Write Header Row (Row 1) ---
    # Original headers from DataFrame map to columns A-G
    headers_map = {
        1: "Project", 2: "Required Supplier", 3: "Manufacturer", 4: "Manufacturer Part #",
        5: "Description", 6: "Quantity", 7: "Cost", 8: "Extended Cost",
        9: "", 10: "Price Expiration" # Add empty col I header and Price Exp header
    }
    for c_idx, header_text in headers_map.items():
        cell = ws.cell(row=1, column=c_idx, value=header_text)
        # Don't apply yellow fill to the empty column I header (c_idx 9)
        if c_idx != 9:
            cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border

    # --- Write Data Rows ---
    last_data_row = 1 # Start after header
    # Use dataframe_to_rows but skip header
    rows = dataframe_to_rows(df, index=False, header=False)

    for r_idx, row_data in enumerate(rows, 2): # Start writing from row 2
        last_data_row = r_idx
        for c_idx, value in enumerate(row_data, 1):
            # We only have data for columns A-G from the DataFrame
            if c_idx > 7: continue # Stop after Cost column data

            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font = normal_font
            cell.border = thin_border

            # Apply specific formatting based on column index
            if c_idx == 6: # Quantity (Column F)
                cell.alignment = center_alignment
                cell.number_format = integer_format
            elif c_idx == 7: # Cost (Column G)
                cell.alignment = right_alignment
                cell.number_format = currency_format
            else: # Columns A-E
                cell.alignment = left_alignment

        # --- Add Formula/Empty Cells for Columns H, I, J ---
        # H: Extended Cost (Formula, Green Fill)
        ext_cost_cell = ws.cell(row=r_idx, column=8)
        ext_cost_cell.value = f"=F{r_idx}*G{r_idx}" # Formula: Quantity * Cost
        ext_cost_cell.fill = ext_cost_fill
        ext_cost_cell.font = normal_font
        ext_cost_cell.alignment = right_alignment
        ext_cost_cell.number_format = currency_format
        ext_cost_cell.border = thin_border

        # I: Empty Column
        empty_i_cell = ws.cell(row=r_idx, column=9, value="")
        empty_i_cell.border = thin_border # Apply border even if empty

        # J: Price Expiration (Empty Data)
        empty_j_cell = ws.cell(row=r_idx, column=10, value="")
        empty_j_cell.border = thin_border # Apply border even if empty

    # --- Add Total Row ---
    if last_data_row > 1: # Only add total if there's data
        total_row = last_data_row + 1

        # Apply border to empty cells A-G in the total row
        for col in range(1, 8):
            cell = ws.cell(row=total_row, column=col)
            cell.border = thin_border

        # H: Sum of Extended Cost (NO GREEN FILL, just borders)
        sum_cell = ws.cell(row=total_row, column=8)
        sum_cell.value = f"=SUM(H2:H{last_data_row})"
        sum_cell.font = normal_font # Normal font for total value
        sum_cell.alignment = right_alignment
        sum_cell.number_format = currency_format
        sum_cell.border = thin_border

        # I: "TOTAL" Label with border
        total_label_cell = ws.cell(row=total_row, column=9, value="TOTAL")
        total_label_cell.font = normal_font # Normal font for "TOTAL" label
        total_label_cell.alignment = right_alignment # RIGHT alignment for TOTAL label
        total_label_cell.border = thin_border

        # J: Empty cell in total row - NO BORDER
        empty_total_j = ws.cell(row=total_row, column=10)

    # --- Save to memory buffer ---
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# --- Streamlit App Layout ---
# Initialize session state for store name and project ID if not already set
if 'store_name' not in st.session_state:
    st.session_state.store_name = ""
    
if 'project_id' not in st.session_state:
    st.session_state.project_id = ""

# Initialize pole quantities dictionary in session state if not already set
if 'pole_quantities_input' not in st.session_state:
    st.session_state.pole_quantities_input = {}
    # Initialize all pole part keys
    for part_num in [k for k, v in clinton_parts.items() if v.get("type") == "pole"]:
        st.session_state.pole_quantities_input[part_num] = 0

# Initialize session state for reader count if not already set
if 'reader_count' not in st.session_state:
    st.session_state.reader_count = 1

# Initialize session state for cable quantity if not already set
if 'cable_quantity' not in st.session_state:
    st.session_state.cable_quantity = 500

# Initialize session state for address input if not already set
if 'address_input' not in st.session_state:
    st.session_state.address_input = ""

# Set up the app layout with minimal file watching
st.set_page_config(page_title="RFID BoM Generator", layout="wide")

# Disable file watcher to avoid inotify limits in deployed environments
if os.environ.get('STREAMLIT_ENV') == 'production' or os.environ.get('STREAMLIT_ENV') == 'cloud':
    st.cache_resource._dont_watch_existing_files = True

st.title("RFID BoM Generator")

# Create tabs
tab1, tab2, tab3, tab4 = st.tabs(["Clinton Pole Order", "Material Order", "Service Now Request", "Help"])

# ----- CLINTON POLE ORDER TAB -----
with tab1:
    st.header("Clinton Pole Order")
    
    # Create a nice card-like container for project information
    with st.container():
        st.subheader("Project Information")
        col1, col2 = st.columns(2)
        
        with col1:
            # Project ID input - using on_change callback instead of direct assignment
            project_id = st.text_input("Project ID", value=st.session_state.project_id, key="project_id_input", 
                                      on_change=lambda: setattr(st.session_state, 'project_id', st.session_state.project_id_input))
            
        with col2:
            # Store name input - using on_change callback instead of direct assignment
            store_name = st.text_input("Store Name", value=st.session_state.store_name, key="store_name_input",
                                      on_change=lambda: setattr(st.session_state, 'store_name', st.session_state.store_name_input))
    
    # Reader count in its own section
    with st.container():
        st.subheader("RFID Reader Count")
        reader_count = st.number_input("Number of RFID Readers", min_value=1, step=1, value=st.session_state.reader_count, key="reader_count_input",
                                     on_change=lambda: setattr(st.session_state, 'reader_count', st.session_state.reader_count_input))
    
    # Add a separator
    st.markdown("---")
    
    # Clinton Pole Selection with improved layout
    st.subheader("Clinton Pole Selection")
    
    # Create a two-column layout for pole selection and summary
    pole_col, summary_col = st.columns([2, 1])
    
    with pole_col:
        # Function to update pole quantities in session state
        def update_pole_qty(part_num):
            # Only update if the key exists in session state
            if f"qty_{part_num}" in st.session_state:
                st.session_state.pole_quantities_input[part_num] = st.session_state[f"qty_{part_num}"]
        
        # Filter the poles from the clinton_parts dictionary
        pole_part_nums = [k for k, v in clinton_parts.items() if v.get("type") == "pole"]
        
        # Initialize the pole quantities dictionary
        pole_quantities_input = {}
        
        # Use expanders for each pole category
        with st.expander("3ft Poles", expanded=True):
            cols = st.columns(2)
            with cols[0]:
                if "CE-CP3W" in clinton_parts:
                    label = "3ft White Pole"
                    pole_quantities_input["CE-CP3W"] = st.number_input(
                        label, 
                        min_value=0, 
                        step=1, 
                        value=st.session_state.pole_quantities_input.get("CE-CP3W", 0), 
                        key=f"qty_CE-CP3W", 
                        on_change=update_pole_qty, 
                        args=("CE-CP3W",)
                    )
            with cols[1]:
                if "CE-CP3B" in clinton_parts:
                    label = "3ft Black Pole"
                    pole_quantities_input["CE-CP3B"] = st.number_input(
                        label, 
                        min_value=0, 
                        step=1, 
                        value=st.session_state.pole_quantities_input.get("CE-CP3B", 0), 
                        key=f"qty_CE-CP3B", 
                        on_change=update_pole_qty, 
                        args=("CE-CP3B",)
                    )
        
        with st.expander("6ft Poles", expanded=True):
            cols = st.columns(2)
            with cols[0]:
                if "CE-CP6W" in clinton_parts:
                    label = "6ft White Pole"
                    pole_quantities_input["CE-CP6W"] = st.number_input(
                        label, 
                        min_value=0, 
                        step=1, 
                        value=st.session_state.pole_quantities_input.get("CE-CP6W", 0), 
                        key=f"qty_CE-CP6W", 
                        on_change=update_pole_qty, 
                        args=("CE-CP6W",)
                    )
            with cols[1]:
                if "CE-CP6B" in clinton_parts:
                    label = "6ft Black Pole"
                    pole_quantities_input["CE-CP6B"] = st.number_input(
                        label, 
                        min_value=0, 
                        step=1, 
                        value=st.session_state.pole_quantities_input.get("CE-CP6B", 0), 
                        key=f"qty_CE-CP6B", 
                        on_change=update_pole_qty, 
                        args=("CE-CP6B",)
                    )

        with st.expander("Adjustable Poles (3'-11\" to 10'-11\")", expanded=True):
            # Single poles layout - using the full width now that we've removed the 2-pack option
            cols = st.columns(2)
            with cols[0]:
                if "CE-CP412W" in clinton_parts:
                    label = "Adjustable White Pole"
                    pole_quantities_input["CE-CP412W"] = st.number_input(
                        label, 
                        min_value=0, 
                        step=1, 
                        value=st.session_state.pole_quantities_input.get("CE-CP412W", 0), 
                        key=f"qty_CE-CP412W", 
                        on_change=update_pole_qty, 
                        args=("CE-CP412W",)
                    )
            with cols[1]:
                if "CE-CP412B" in clinton_parts:
                    label = "Adjustable Black Pole"
                    pole_quantities_input["CE-CP412B"] = st.number_input(
                        label, 
                        min_value=0, 
                        step=1, 
                        value=st.session_state.pole_quantities_input.get("CE-CP412B", 0), 
                        key=f"qty_CE-CP412B", 
                        on_change=update_pole_qty, 
                        args=("CE-CP412B",)
                    )
            
            # Add info message about 2-pack optimization
            st.info("**Note:** When ordering 2 or more adjustable poles, the system will automatically use 2-packs to optimize cost.")
            
            # 2-Pack poles are now handled automatically and not shown in the UI
            # We still add them to the dictionary with zero quantity for the algorithm to use
            if "CE-CP412W-2PK" in clinton_parts:
                pole_quantities_input["CE-CP412W-2PK"] = 0
            if "CE-CP412B-2PK" in clinton_parts:
                pole_quantities_input["CE-CP412B-2PK"] = 0
        
        # Add other poles not in the main categories directly to pole_quantities_input 
        # but don't display them in the UI
        other_poles = [p for p in pole_part_nums if p not in ["CE-CP3W", "CE-CP3B", "CE-CP6W", "CE-CP6B", "CE-CP412B", "CE-CP412B-2PK", "CE-CP412W", "CE-CP412W-2PK"]]
        if other_poles:
            for part_num in other_poles:
                # Silently add them to the dictionary with zero quantity
                pole_quantities_input[part_num] = 0
    
    # Order Summary column
    with summary_col:
        with st.container():
            st.markdown("### Order Summary")
            
            # Calculate total poles selected
            total_poles = 0
            selected_poles = []
            
            for part_num, qty in pole_quantities_input.items():
                if qty > 0:
                    # Count 2-packs as 2 poles each
                    if "-2PK" in part_num:
                        actual_qty = qty * 2
                        total_poles += actual_qty
                        selected_poles.append({
                            "name": clinton_parts[part_num]['desc'],
                            "qty": qty,
                            "actual_qty": actual_qty
                        })
                    else:
                        total_poles += qty
                        selected_poles.append({
                            "name": clinton_parts[part_num]['desc'],
                            "qty": qty,
                            "actual_qty": qty
                        })
            
            # Display total pole count
            st.metric("Total Poles", total_poles)
            
            # Display selected poles
            if selected_poles:
                st.markdown("#### Selected Poles")
                for pole in selected_poles:
                    if "-2PK" in pole["name"]:
                        st.markdown(f"• {pole['qty']} x {pole['name']} ({pole['actual_qty']} poles total)")
                    else:
                        st.markdown(f"• {pole['qty']} x {pole['name']}")
            else:
                st.info("No poles selected yet")
            
            # Calculate and display mounting accessories needed
            if total_poles > 0:
                st.markdown("#### Accessories (Auto-Added)")
                st.markdown(f"• {total_poles} x Universal Mounting Plates (based on pole count)")
                st.markdown(f"• {total_poles} x Beam Clamps (based on pole count)")
    
    # Add a separator
    st.markdown("---")
    
    # BoM Generation Section
    st.subheader("Generate Bill of Materials")
    
    # Help text
    st.info("Click the button below to generate a Bill of Materials for the selected poles and accessories.")
    
    # Generate BoM Button
    if st.button("Generate Clinton BoM"):
        if project_id and store_name:
            # Update session state
            st.session_state.pole_quantities_input = pole_quantities_input
            
            # Generate BoM
            clinton_bom = generate_clinton_bom(project_id, store_name, reader_count, pole_quantities_input)
            
            if clinton_bom:
                # Create a DataFrame from the BoM
                df = pd.DataFrame(clinton_bom)
                
                # Add Extended Cost column calculation
                if "Cost" in df.columns and "Quantity" in df.columns:
                    df["Extended Cost"] = df["Cost"] * df["Quantity"]
                
                # Calculate total cost
                total_cost = df["Extended Cost"].sum() if "Extended Cost" in df.columns else 0
                
                # Display the BoM
                st.success(f"Bill of Materials Generated - Total Cost: ${total_cost:.2f}")
                st.dataframe(df)
                
                # Create Excel file
                excel_bytes = create_excel_bytes(df, f"{project_id} - {store_name} - Clinton BoM")
                
                # Download button
                st.download_button(
                    label="Download Clinton BoM Excel",
                    data=excel_bytes,
                    file_name=f"{project_id}_{store_name}_Clinton_BoM.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            else:
                st.error("No items in the Bill of Materials. Please select at least one pole.")
        else:
            st.error("Please enter Project ID and Store Name to generate a BoM.")

# ----- MATERIAL ORDER TAB -----
with tab2:
    st.header("Material Order")

    # --- Input Section ---
    st.subheader("1. Enter Material Order Details")

    # Use columns for better layout
    col1, col2 = st.columns([1, 2])

    with col1:
        # Display the shared project ID, store name, reader count from session state
        st.write(f"**Project ID:** {st.session_state.project_id}")
        st.write(f"**Store Name:** {st.session_state.store_name}")
        st.write(f"**Reader Count:** {st.session_state.reader_count}")

        # Additional input for cable quantity specific to this tab
        def update_cable_qty():
            st.session_state.cable_quantity = st.session_state.cable_qty_input

        cable_quantity_input = st.number_input("Network Cable Boxes (1000' each):", min_value=0, step=1,
                                         value=st.session_state.cable_quantity, key="cable_qty_input",
                                         on_change=update_cable_qty)

    with col2:
        st.subheader("Material Quantities (Calculated)")
        st.write("Quantities based on Reader Count and Cable Boxes:")
        # Calculate derived quantities for display
        reader_count = st.session_state.reader_count
        jacks_qty = reader_count * 2
        smb_qty = reader_count
        patch1_qty = reader_count # Assuming 10' patch cord is also per reader

        st.markdown(f"- **White CMP CAT6 Cable (10136230):** {st.session_state.cable_quantity} (Input Above)")
        st.markdown(f"- **Blue Cat6 Jack (NK688MBU):** {jacks_qty} ({reader_count} Readers × 2)")
        st.markdown(f"- **2 Port SMB (NK2BXWH-A):** {smb_qty} ({reader_count} Readers)")
        st.markdown(f"- **1' White Cat6 Patch Cord (INFINI CAB CAT6-01WH):** {patch1_qty} ({reader_count} Readers)")
        st.markdown(f"- **10' White Cat6 Patch Cord (AT1610-WH):** {patch1_qty} ({reader_count} Readers)")
        st.markdown("- **All other materials:** Quantity of 1")

    # --- Generate Material BoM Button and Output Section ---
    st.header("2. Generate Material Bill of Materials")

    if st.button("Generate Material BoM", key="generate_material_button"):
        # Basic validation
        if not st.session_state.project_id:
            st.warning("⚠️ Please enter a Project ID in the Clinton Pole Order tab.")
            st.stop()

        if st.session_state.reader_count <= 0 and st.session_state.cable_quantity <= 0 :
             st.warning("⚠️ Please enter a Reader Count (> 0) in the Clinton Pole Order tab or Cable Box quantity (> 0).")
             st.stop()

        # Call the generation function using current session state values
        generated_bom_list = generate_material_bom(st.session_state.project_id,
                                                  st.session_state.reader_count,
                                                  st.session_state.cable_quantity) # Use updated cable qty

        if generated_bom_list:
            # Convert list of dictionaries to Pandas DataFrame
            bom_df = pd.DataFrame(generated_bom_list)

            # Define the exact column order needed for create_excel_bytes (A-G)
            column_order = ["Project", "Required Supplier", "Manufacturer", "Manufacturer Part #",
                            "Description", "Quantity", "Cost"]
            # Ensure all expected columns exist, add if missing
            for col in column_order:
                if col not in bom_df.columns:
                    bom_df[col] = None
            bom_df = bom_df[column_order]  # Reorder

            # --- Display in Streamlit (Optional - formatted) ---
            display_df = bom_df.copy()
            display_df["Extended Cost"] = display_df["Quantity"] * display_df["Cost"]
            display_df["Cost"] = display_df["Cost"].map('${:,.2f}'.format)
            display_df["Extended Cost"] = display_df["Extended Cost"].map('${:,.2f}'.format)
            display_df[" "] = "" # Column I placeholder
            display_df["Price Expiration"] = "" # Column J placeholder
            display_order = ["Project", "Required Supplier", "Manufacturer", "Manufacturer Part #",
                             "Description", "Quantity", "Cost", "Extended Cost", " ", "Price Expiration"]
            st.subheader("Generated Material BoM (Preview)")
            st.dataframe(display_df[display_order], hide_index=True, use_container_width=True)

            # --- Export to Excel ---
            # Create the Excel file using the consolidated function
            excel_bytes = create_excel_bytes(bom_df[column_order], "Material BoM")

            # Determine filename based on store name
            filename = f'RFID Material Order {st.session_state.store_name}'
            if not st.session_state.store_name:
                filename = f'RFID Material Order {st.session_state.project_id.replace(" ", "_").replace("/", "-")}'

            # Download button
            st.download_button(
                label="Download Material BoM as Excel",
                data=excel_bytes,
                file_name=f'{filename}.xlsx',
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                key='download_material_excel'
            )

            # Store raw data in session state
            st.session_state['generated_material_bom_df_raw'] = bom_df
        else:
            st.info("ℹ️ No Material BoM items were generated.")
            if 'generated_material_bom_df_raw' in st.session_state:
                 del st.session_state['generated_material_bom_df_raw']

# ----- SERVICE NOW REQUEST TAB -----
with tab3:
    st.header("Service Now Procurement Request")

    # Load site data from CSV
    site_data = None
    csv_path = "pm_project_task.csv"
    
    if os.path.exists(csv_path):
        try:
            site_data = pd.read_csv(csv_path)
            # Clean up the Site ID column to extract just the store number for filtering
            site_data['Site Number'] = site_data['Site ID'].apply(lambda x: x.split(' - ')[0] if ' - ' in str(x) else x)
            st.success(f"Successfully loaded {len(site_data)} site locations from {csv_path}")
        except Exception as e:
            st.error(f"Error loading site data: {e}")
    else:
        st.warning(f"Site data file not found: {csv_path}")

    # Project information
    sn_col1, sn_col2 = st.columns(2)
    
    with sn_col1:
        # Use the existing project_id and store_name values from other tabs
        project_id = st.text_input("Project ID", value=st.session_state.project_id, key="sn_project_id")
        
        # Store name with site lookup - use existing store_name from session state
        store_name = st.text_input("Store Name", value=st.session_state.store_name, key="sn_store_name")
        
        # Function to find the best matching site using fuzzy matching
        def find_best_matching_site(query, sites, score_cutoff=70):
            """
            Find the best matching site in the site list using fuzzy string matching.
            
            Args:
                query (str): The store name to match
                sites (list): List of site IDs to match against
                score_cutoff (int): Minimum score (0-100) to consider a match
                
            Returns:
                str: The best matching site ID or None if no good match
            """
            if not query or not sites:
                return None
                
            # Try to match the query with the site IDs
            best_match = process.extractOne(query, sites)
            if best_match and best_match[1] >= score_cutoff:
                return best_match[0]
            return None
        
        # Add a Site ID selection dropdown if the CSV was loaded
        if site_data is not None:
            # If store name is entered, try to find a matching site with fuzzy search
            matched_site = None
            if store_name:
                # Clean the store name for better matching (focusing on store number)
                clean_store_name = store_name.strip().upper()
                
                # Try to extract just the store number pattern (e.g., "GRG 387")
                store_pattern = re.search(r'([A-Z]+)\s*(\d+)', clean_store_name)
                if store_pattern:
                    store_code = store_pattern.group(1)
                    store_number = store_pattern.group(2)
                    search_pattern = f"{store_code} {store_number}"
                    matched_site = find_best_matching_site(search_pattern, site_data['Site ID'].tolist())
                else:
                    # If no clear pattern, try matching the whole store name
                    matched_site = find_best_matching_site(clean_store_name, site_data['Site ID'].tolist())
                
                if matched_site:
                    st.success(f"Found matching site: {matched_site}")
            
            st.subheader("Site Lookup")
            
            # Create a searchable dropdown for site selection
            site_options = [""] + site_data['Site ID'].tolist()
            
            # If we found a matching site, pre-select it in the dropdown
            default_index = 0
            if matched_site and matched_site in site_options:
                default_index = site_options.index(matched_site)
                
            selected_site = st.selectbox(
                "Select a site to auto-fill address information",
                options=site_options,
                index=default_index,
                format_func=lambda x: x if x else "Select a site..."
            )
            
            # Auto-fill the address when a site is selected
            if selected_site:
                # Find the selected site in the dataframe
                site_row = site_data[site_data['Site ID'] == selected_site].iloc[0]
                
                # Format the complete address for display
                address1 = site_row['Address 1'] if not pd.isna(site_row['Address 1']) else ""
                address2 = site_row['Address 2'] if not pd.isna(site_row['Address 2']) else ""
                city = site_row['City'] if not pd.isna(site_row['City']) else ""
                state = site_row['State / Province'] if not pd.isna(site_row['State / Province']) else ""
                zip_code = site_row['Zip / Postal Code'] if not pd.isna(site_row['Zip / Postal Code']) else ""
                country = site_row['Country'] if not pd.isna(site_row['Country']) else ""
                
                full_address = f"{address1}"
                if address2:
                    full_address += f"\n{address2}"
                full_address += f"\n{city}, {state} {zip_code}"
                if country:
                    full_address += f"\n{country}"
                
                # Set the address in the session state
                st.session_state.address_input = full_address
        
        material_arrival_date = st.date_input("Material Arrival Date", key="arrival_date")
        
        # Parse the address - allow for a multi-line input
        st.subheader("Store Address")
        address_input = st.text_area(
            "Enter the full address (will be parsed automatically)",
            height=100,
            key="address_input"
        )
        
        # Simple address parser
        address_lines = address_input.strip().split('\n')
        parsed_address = {
            "street": address_lines[0] if len(address_lines) > 0 else "",
            "city_state_zip": address_lines[1] if len(address_lines) > 1 else "",
            "additional": "\n".join(address_lines[2:]) if len(address_lines) > 2 else ""
        }
        
    with sn_col2:
        # Display parsed address components
        if address_input.strip():
            st.subheader("Parsed Address")
            st.write("**Street:** " + parsed_address["street"])
            st.write("**City, State, ZIP:** " + parsed_address["city_state_zip"])
            if parsed_address["additional"]:
                st.write("**Additional Info:**")
                st.write(parsed_address["additional"])
    
    # Create the procurement request text
    if st.button("Generate Service Now Requests"):
        if not all([project_id, store_name, address_input.strip()]):
            st.error("Please fill in all the required fields.")
        else:
            # Format store name for GRG stores
            display_store_name = store_name
            if "GRG" in store_name.upper():
                # Extract the number from GRG store name
                import re
                grg_number = re.search(r'\d+', store_name)
                if grg_number:
                    display_store_name = f"GARAGE CLOTHING {grg_number.group()}"
                else:
                    display_store_name = f"GARAGE CLOTHING"
            
            # Format the date as MM/DD/YYYY
            formatted_date = material_arrival_date.strftime("%m/%d/%Y")
            
            # Generate request for Anixter
            st.subheader("Anixter Procurement Request")
            anixter_request = f"""Please order the following per attached BOM through Anixter to be delivered to the following address on {formatted_date}:

{display_store_name}
ATTN: HOLD FOR TELAID
{parsed_address["street"]}
{parsed_address["city_state_zip"]}
{parsed_address["additional"]}

Telaid Contact: gdrfid@telaid.com

Special Delivery Instructions: Inside Delivery Required"""
            
            st.text_area("Anixter Request Text (Copy this to Service Now)", anixter_request, height=300)
            
            # Generate request for Clinton
            st.subheader("Clinton Procurement Request")
            clinton_request = f"""Please order the following per attached BOM through Clinton to be delivered to the following address on {formatted_date}:

{display_store_name}
ATTN: HOLD FOR TELAID
{parsed_address["street"]}
{parsed_address["city_state_zip"]}
{parsed_address["additional"]}

Telaid Contact: gdrfid@telaid.com

Special Delivery Instructions: Inside Delivery Required"""
            
            st.text_area("Clinton Request Text (Copy this to Service Now)", clinton_request, height=300)

# ----- HELP TAB -----
with tab4:
    st.header("Help")
    
    # How to Run section
    st.subheader("How to Run")
    
    col1, col2 = st.columns([1, 2])
    
    with col1:
        st.markdown("**1. Save:**")
        st.markdown("**2. Requirements:**")
        st.markdown("**3. Run:**")
    
    with col2:
        st.markdown("Save this code as `rfid_bom_generator.py`.")
        st.markdown("Ensure `requirements.txt` contains `streamlit`, `pandas`, and `openpyxl`. Install with `pip install -r requirements.txt`.")
        st.markdown("Use the `start_rfid_bom.bat` file or run from terminal: `streamlit run rfid_bom_generator.py`")
    
    # Adding More Parts section
    st.subheader("Add More Parts")
    
    # Clinton Parts
    st.markdown("**Clinton Poles/Accessories:**")
    st.markdown("Edit the `clinton_parts = {...}` dictionary.")
    st.code("""Format: "PART-NUM": {"desc": "Desc", "type": "pole" or "accessory", "cost": price}""", language="python")
    
    # Material Parts
    st.markdown("**Material Parts:**")
    st.markdown("Edit the `material_parts = {...}` dictionary.")
    st.code("""Format: "PART-NUM": {"desc": "Desc", "manufacturer": "Mfg", "supplier": "Supplier", "cost": price}""", language="python")
    
    # Using the Application section
    st.subheader("Using the Application")
    
    st.markdown("""
    **Clinton Pole Order Tab:**
    1. Enter the Project ID and Store Name
    2. Set the number of RFID Readers
    3. Select the quantities for each pole type
    4. Click 'Generate Clinton BoM' to create the Bill of Materials
    5. You can download the Excel file with the formatted BoM
    
    **Material Order Tab:**
    1. Project ID and Store Name are shared from the Clinton tab
    2. Adjust Reader Count if needed
    3. Set the White CMP CAT6 Cable quantity
    4. Click 'Generate Material BoM' to create the material list
    5. Download the Excel file with the complete BoM
    
    **Service Now Request Tab:**
    1. Auto-fills Project ID and Store Name from previous tabs
    2. Provides a store lookup feature using the site database
    3. Fuzzy search automatically finds matching sites
    4. Enter or select an address and parse it automatically
    5. Generate procurement requests for Anixter and Clinton
    """)
    
    # 2-Pack Optimization note
    st.info("""
    **Note:** When ordering 2 or more of the CE-CP412B (Black) or CE-CP412W (White) adjustable poles, 
    the system will automatically optimize your order to use the 2-pack option (CE-CP412B-2PK or CE-CP412W-2PK), 
    saving $7.68 per pair. You can also manually select the 2-pack options if desired.
    """)

if __name__ == "__main__":
    # Ensure all initialization code is above this point
    import asyncio
    
    # Handle potential asyncio issues in different environments
    try:
        asyncio.set_event_loop_policy(asyncio.DefaultEventLoopPolicy())
    except RuntimeError:
        pass  # Ignore if event loop is already running